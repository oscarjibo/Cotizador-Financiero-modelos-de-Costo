# -*- coding: utf-8 -*-
"""
Created on Wed Jul  6 18:58:05 2022

@author: oscar
"""
import main
from main import*
import main as mr
from datetime import date
from datetime import datetime
from PyQt5.QtWidgets import QApplication, QWidget, QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, \
                            QPushButton, QItemDelegate, QVBoxLayout
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QDoubleValidator
# voy ac rear el dataframe 

import openpyxl


class FloatDelegate(QItemDelegate):
    def __init__(self, parent=None):
        super().__init__()

    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setValidator(QDoubleValidator())
        return editor

class TableWidget(QTableWidget):
    def __init__(self, df):
        super().__init__()
        self.df = df
        self.setStyleSheet('font-size: 20px;')

        # set table dimension
        nRows, nColumns = self.df.shape
        self.setColumnCount(nColumns)
        self.setRowCount(nRows)

        self.setHorizontalHeaderLabels(('Articulo', 'Cantidad','Costo','Costo Mensual','Tipo','Unitario'))

        self.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.setItemDelegateForColumn(1, FloatDelegate())

        # data insertion
        for i in range(self.rowCount()):
            for j in range(self.columnCount()):
                self.setItem(i, j, QTableWidgetItem(str(self.df.iloc[i, j])))

        self.cellChanged[int, int].connect(self.updateDF)   

    def updateDF(self, row, column):
        text = self.item(row, column).text()
        self.df.iloc[row, column] = text

class DFEditor(QWidget):
    
    # abrimos el excel como un dataframe


    f = open ('file.txt','r')
    mensaje = f.read()

    f.close()

    df = pd.read_excel(mensaje)
    df=df.round({'Costo': 1})
    df=df.round({'Costo mensual': 1})
    df['Costo']=df['Costo'].apply('{:,}'.format)
    #df['Costo mensual']=df['Costo mensual'].apply('{:,}'.format)

    G=df.to_numpy().transpose().tolist()

    lista1 =G[4]
    lista2=G[2]
    lista3=[]
    for i in range(0,len(lista1)):
        lista3.append(lista1[i]//lista2[i])


    data = {
        'Articulo': G[1],
        'Cantidad': G[2],
        'Costo': G[3],
        'Costo Mensual': G[4],
        'Tipo': G[5],
        'Unitario':lista3
        
    }

    df = pd.DataFrame(data)

    def __init__(self):
        super().__init__()
        self.resize(1500, 720)

        mainLayout = QVBoxLayout()

        self.table = TableWidget(DFEditor.df)
        mainLayout.addWidget(self.table)

        button_print = QPushButton('Guardar como nueva Version')
        button_print.setStyleSheet('font-size: 20px')
        button_print.clicked.connect(self.print_DF_Values)
        mainLayout.addWidget(button_print)

        button_export = QPushButton('regresar')
        button_export.setStyleSheet('font-size:20px')
        button_export.clicked.connect(self.export_to_csv)
        mainLayout.addWidget(button_export)     
        

        button_export = QPushButton('añadir articulos')
        button_export.setStyleSheet('font-size:20px')
        button_export.clicked.connect(self.agregar)
        mainLayout.addWidget(button_export)    

        self.setLayout(mainLayout)
    def agregar(self):
        self.hide()
        otraventana=agregarM(self)
        otraventana.show()
    def print_DF_Values(self):
        print(self.df)
        self.df=self.df
        a=0#capex
        b=0#opex1t
        c=0#otropex
        d=0#personal
        tipe=(self.df['Tipo'])
        unit=(self.df['Unitario'])
        cant=(self.df['Cantidad'])
        for i in range(len(self.df['Tipo'])):
            aaa=(self.df['Tipo'][i])
            if aaa=='capex':
                aab=float(unit[i])*float(cant[i])
                a+=aab
            if aaa=='opex1t':
                aab=float(unit[i])*float(cant[i])
                b+=aab
            if aaa=='otroopex':
                aab=float(unit[i])*float(cant[i])
                c+=aab
            if aaa=='persona':
                aab=float(unit[i])*float(cant[i])
                d+=aab

        # traigo los datos que neceisto del excel 
        
        print(a,b,c,d)
        #totales
        v=1
        f1=self.df
        f = open ('file.txt','r')
        mensaje = f.read()
        print(mensaje)
        f.close()
        mensaje=mensaje.replace(".xlsx","")
        print(mensaje)
        file_name1 =mensaje + 'v1'+ '.xlsx'
        mensaje1=mensaje

        f1.to_excel(file_name1)
        df = pd.read_excel(archivo_reportes)
        
        
        
        #abrir excel con el registro
        mensaje=mensaje+'reporte'+'.xlsx'
        excel_document = openpyxl.load_workbook(mensaje)
          
        sheet_obj = excel_document.active 
          
          
          
        cell_obj1 = sheet_obj.cell(row = 11, column = 2) 
        cell_obj2 = sheet_obj.cell(row = 10, column = 2) 
        cell_obj3 = sheet_obj.cell(row = 12, column = 2) 
        cell_obj4 = sheet_obj.cell(row = 6, column = 2) 
        cell_obj5 = sheet_obj.cell(row = 20, column = 2) 


        cell_obj7 = sheet_obj.cell(row = 19, column = 5) 
        cell_obj8 = sheet_obj.cell(row = 20, column = 5) 
        cell_obj9 = sheet_obj.cell(row = 21, column = 5) 
        cell_obj10 = sheet_obj.cell(row = 18, column = 5)
        
        cell_obj11 = sheet_obj.cell(row = 18, column = 2) 
        cell_obj12 = sheet_obj.cell(row = 19, column = 2) 
        cell_obj13 = sheet_obj.cell(row = 18, column = 5) 

        


        comercial= cell_obj1.value
        preventa=cell_obj2.value
        cliente=cell_obj3.value
        TOTAL=cell_obj4.value
        CRM=cell_obj5.value
        fecha=today
        tipoofer="new version"
        moneda="cop"
        tasa_capex=cell_obj7.value
        tasa_opex1t=cell_obj8.value
        tasa_otroopex=cell_obj9.value
        tasa_personal=cell_obj10.value
        riesgo=cell_obj11.value
        agregados=cell_obj12.value
        ipc=1
        mesesfact=36
        
        
        
        
        nuevo_registro = {'comercial': comercial, 'preventa': preventa, 'cliente': cliente,
                          'TOTAL': TOTAL, 'CRM': CRM,'fecha':today,"tipo" : tipoofer, 'Moneda':moneda}
        df = df.append(nuevo_registro, ignore_index=True)
        file_name = 'registros.xlsx'
        df.to_excel(file_name, index=False)
        a1=(a/(1-(riesgo/100)-(tasa_capex/100)-(agregados/100)))*(1+(ipc/100))
        b1=(b/(1-(riesgo/100)-(tasa_opex1t/100)-(agregados/100)))*(1+(ipc/100))
        c1=(c/(1-(riesgo/100)-(tasa_otroopex/100)-(agregados/100)))*(1+(ipc/100))
        d1=(d/(1-(riesgo/100)-(tasa_personal/100)-(agregados/100)))*(1+(ipc/100))
        # abro el arcivho ene xcel del reporte

        print(a1+b1+c1+d1)
        # voy a crear el nuevo archivo excel 
        wb = load_workbook(filesheet)
        sheet = wb.active
        sheet['A4'] = "VALOR Mensual"
        # Ingresamos el valor 1845 en la celda 'B3'
        sheet['B2']=today
        sheet['B3'] = "Capex"
        sheet['C3'] = "Otro opex"
        sheet['D3'] = "Opex 1t"
        sheet['E3'] = "Personal"
        sheet['B4'] = a1
        sheet['C4'] = c1
        sheet['D4'] = b1
        sheet['E4'] = d1
        sheet['A5'] = "TOTAL mensual"
        sheet['B5'] = (a1+b1+c1+d1)
        sheet['A6'] = "TOTAL PROYECTO"
        total2=(a1+b1+c1+d1)*mesesfact
        sheet['B6'] = total2
        sheet['A10'] = "Preventa"
        sheet['B10'] = preventa
        sheet['A11'] = "Comercial"
        sheet['B11'] = comercial
        sheet['B17'] = ipc
        sheet['B18'] = riesgo
        sheet['B19'] = agregados
        sheet['B20'] = CRM
        sheet['B12'] = cliente
        sheet['B2'] = today
        file_name2 = mensaje1+'reporte'+'V1'+'.xlsx'

       # os.startfile(file_name1)
        # Guardamos el archivo con los cambios
        wb.save(file_name2)
        os.startfile(file_name2) 
        

    def export_to_csv(self):
        self.hide()
        otraventana=Modificar(self)
        otraventana.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)

    demo = DFEditor()
    demo.show()
    
    sys.exit(app.exec_())