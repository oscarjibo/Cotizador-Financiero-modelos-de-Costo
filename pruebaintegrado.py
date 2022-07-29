# -*- coding: utf-8 -*-
"""
Created on Tue Jul 12 09:07:24 2022

@author: oscar
"""
from PyQt5.QtWidgets import QApplication, QWidget, QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit, \
                            QPushButton, QItemDelegate, QVBoxLayout
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QDoubleValidator
# voy ac rear el dataframe 

import tkinter as tk
from pandastable import Table, TableModel
import numpy as np
import numpy_financial as npf
import pathlib
import os
from PyQt5 import QtCore
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QTableView)
from PyQt5.QtCore import (QAbstractTableModel, Qt)
import sys
from PyQt5 import uic, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow,QLCDNumber
from PyQt5.uic import loadUi
qtCreatorFile = "principal.ui" 
# para crear el PDF
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
w, h = A4
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
from datetime import datetime

today = date.today()
# ruta de nuestros archivos
filesheet = ".\ejercicio.xlsx"
miarchivo =".\ejercicio.xlsx"
archivo_adicion=".\\adicion.docx"
archivo_salarios=".\\salarios.xlsx"
archivo_nuevo=".\\nuevo.docx"
archivo_contactos=".\contactos.xlsx"
archivo_usuarios=".\\usuarios.xlsx"
archivo_reportes=".\\registros.xlsx"
archivo_comerciales=".\Directorio.pdf"
archivousd=".\mainusd.py"
mipdf=".\oferta.pdf"


# lista para el combobox
registros=archivo_reportes
reg = pd.read_excel(registros)
reg1=reg.loc[:,'cliente']
reg1=list(reg1)
print(type(reg1))
print(reg1)
#datos nombres
preventa=[]
comercial=[]
cliente=[]
fecha=date.today()
total=[]
# DATOS
moneda=[]
cliente=[] 
tipoofer=[]
tasa_financiacion_general=[]
tasa_capex=[]
tasa_opex1t=[]
tasa_otroopex=[]
tasa_personal=[]
mesesfact=[]
TRM=[]
TRM_SERV=[]
TRM2=[]
riesgototal=[]
imprevistos_1=[]
ipc=[]
crm=[]
linea=["MDS"]
#empiezo a crear mi dataframe 
#TOTALEs
total_personal=[0]
total_otroopex=[0]
total_opex1t=[0]
total_capex=[0]
# datos para mi datafrma
costo_producto=[]
articulos=[]
costos=[]
cantidades=[]
tipo=[]
seguro=[]
componente=[]
componente_status=[]
# meter el dataframe de los sueldos 
version="1"
nm=[]
m = archivo_salarios
df = pd.read_excel(m)
Analista1=df.iloc [0, 7]
Analista2=df.iloc [1, 7]
Analista3=df.iloc [2, 7]
Profesional1=df.iloc [3, 7]
Profesional2=df.iloc [4, 7]
Profesional3=df.iloc [5, 7]
Profesional4=df.iloc [6, 7]
Profesional5=df.iloc [7, 7]
Profesional6=df.iloc [8, 7]
Profesional7=df.iloc [9, 7]
Profesional8=df.iloc [10, 7]
Experto=df.iloc [11, 7]
En1=651900
En2=296150
En3=647500
En4=291750
En5=225800
En6=581550
En7=392630
En8=196315
En9=130877
En10=80550

Ui_MainWindow, QtBaseClass = uic.loadUiType(qtCreatorFile)

class Ventana1(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        QtWidgets.QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)
        self.boton_iniciar.clicked.connect(self.Variables) # abro la ventana de variables
        self.boton_salir.clicked.connect(self.close)# boton salir 
        self.boton_iniciar_2.clicked.connect(self.admin) # abro la ventana de variables

   
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
         pre=self.lista_gente.currentText()
         pcom=self.lista_gente_2.currentText()
         preventa.append(pre)
         comercial.append(pcom)
    def admin(self):
        self.hide()
        otraventana=Admin(self)
        otraventana.show()        
class Admin(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 
    def __init__(self,parent=None):
        super(Admin, self).__init__(parent)
        loadUi('admin.ui', self)
        self.boton_4.clicked.connect(self.close)
        self.boton_3.clicked.connect(self.usuarios)
        self.boton_5.clicked.connect(self.contactos)
        self.boton_6.clicked.connect(self.comerciales)
        self.boton.clicked.connect(self.reportes)
        self.boton_7.clicked.connect(self.salarios)
        
        
    
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
    def usuarios(self):
        os.startfile(archivo_usuarios)
    def reportes(self):
        os.startfile(archivo_reportes)
    def contactos(self):
        os.startfile(archivo_contactos)
    def comerciales(self):
        os.startfile(archivo_comerciales)
    def salarios(self):
        os.startfile(archivo_salarios)         
class Personal(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase del personal
 
    def __init__(self,parent=None):
        super(Personal, self).__init__(parent)
        loadUi('Personal.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre agregar
        self.boton_regresar1.clicked.connect(self.anadir)#boton abre agregar

    
    def agregar(self):
        self.hide()
        otraventana=agregar(self)
        otraventana.show() 
    def anadir(self):
        rol=self.c1.toPlainText()
        perfil=self.c3.currentText()
        meses_ejecutados=self.c2.value()
        cantidad=self.c4.value()
        recargos=self.c5.value()
        ubicacion=self.c6.currentText()
        mesesno=self.c22.value()
        costo=1
        if perfil=="Analista":
            costo=Analista1
        elif perfil=="Analista +":
            costo=Analista2
        elif perfil=="Analista ++":
            costo=Analista3
        elif perfil=="Profesional Junior":
            costo=Profesional1
        elif perfil=="Profesional Pleno":
            costo=Profesional2
        elif perfil=="Profesional Pleno +":
            costo=Profesional3
        elif perfil=="Profesional Especialista ":
            costo=Profesional4
        elif perfil=="Profesional Especialista +":
            costo=Profesional5
        elif perfil=="Profesional Especialista ++":
            costo=Profesional6
        elif perfil=="Profesional Senior ":
            costo=Profesional7
        elif perfil=="Profesional Senior +":
            costo=Profesional8
        elif perfil=="Experto ":
            costo=Experto
        if ubicacion=="En Sonda Pc":
            ubi=En1
        elif ubicacion=="En Cliente Pc":
            ubi=En2
        elif ubicacion=="SONDA Portatil":
            ubi=En3
        elif ubicacion=="Cliente Portatil":
            ubi=En4
        elif ubicacion=="En cliente sin equipo ":
            ubi=En5
        elif ubicacion=="En Sonda sin equipo":
            ubi=En6
        elif ubicacion=="Zona Franca RU1":
            ubi=En7
        elif ubicacion=="Zona Franca RU2":
            ubi=En8
        elif ubicacion=="Zona Franca RU3":
            ubi=En9
        elif ubicacion=="Field Services EUS":
            ubi=En10
        P=(costo)+ubi
        cuota=P*cantidad*(meses_ejecutados+mesesno)/mesesfact[0]
        cuota_total_personal_=npf.pmt(tasa_financiacion_general[0]/100,mesesfact[0]/(meses_ejecutados+mesesno),-1*P*cantidad,0,1)
        total_personal.append(cuota_total_personal_)
        articulos.append(rol)
        cantidades.append(cantidad)
        costos.append(cuota_total_personal_)
        costo_producto.append(P)
        tipo.append("personal")# aca empiezo a darle con los precios del personal 
        seguro.append("no")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class Otroopex(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de otro opex
 
    def __init__(self,parent=None):
        super(Otroopex, self).__init__(parent)
        loadUi('Otroopex.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre personal
        self.boton_anadir.clicked.connect(self.anadir)
        
    
    def agregar(self):
        self.hide()
        otraventana=agregar(self)
        otraventana.show()
    def anadir(self):
    
        item=self.textEdit.toPlainText()
        meses_ejecutado=self.spinBox.value()
        cantidad=self.c.value()
        costo=self.c1.value()
        costo_producto.append(costo)
        iva=self.i.value()
        trm_servicio=self.c3.isChecked()
        if trm_servicio==True:
            costo=TRM_SERV[0]
        unit=costo*(1+iva/100)
        
        cuota_mensual_otropex=npf.pmt(tasa_financiacion_general[0]/100,mesesfact[0]/meses_ejecutado,-1*cantidad*unit,0,1)
        total_otroopex.append(cuota_mensual_otropex)
        articulos.append(item)
        cantidades.append(cantidad)
        costos.append(cuota_mensual_otropex)
        tipo.append("otroopex")
        seguro.append("no")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class opex1t(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 
 
    def __init__(self,parent=None):
        super(opex1t, self).__init__(parent)
        loadUi('opex1t.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre personal
        self.anadi.clicked.connect(self.anadir)
    def agregar(self):
        self.hide()
        otraventana=agregar(self)
        otraventana.show()
    def anadir(self):
    
        item=self.textEdit.toPlainText()
        veces_ejecutado=self.v.value()
        cantidad=self.c.value()
        nacio=self.n.value()
        costo=self.c1.value()
        costo_producto.append(costo)
        iva=self.i.value()
        unitario= costo*(1+nacio/100)*(1+iva/100)

        cuota_mensual_opex1t=npf.pmt(tasa_financiacion_general[0]/100,mesesfact[0]/veces_ejecutado,-1*unitario*cantidad,0,1)
        total_opex1t.append(cuota_mensual_opex1t)
        articulos.append(item)
        cantidades.append(cantidad)
        costos.append(cuota_mensual_opex1t)
        tipo.append("opex1t")
        seguro.append("no")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class Capex(QtWidgets.QMainWindow, Ui_MainWindow):# esta es la clase de Opex 1t 
         
    def __init__(self,parent=None):
        super(Capex, self).__init__(parent)
        loadUi('Capex.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre personal
        self.boton_anadir.clicked.connect(self.anadir)#boton añade
        
    
    def agregar(self):
        self.hide()
        otraventana=agregar(self)
        otraventana.show()
    def anadir(self):
        
        print("añadiendo")
        # aca empiezo a darle con los precios 
        item=self.c1.toPlainText()
        empresa_fabricante=self.q1.toPlainText()
        mesesnofact=self.c2.value()
        nacionalizacion=self.c3.value()
        plataforma=self.c4.value()
        iva=self.c5.value()
        tasa_financiacion=self.c6.value()
        cantidad=self.c7.value()
        costo=self.c8.value()
        costo_producto.append(costo)
        dolar=self.c9.isChecked()
        seguro1=self.c10.isChecked()
        if dolar==True:
            costo=costo*TRM[0]
        if seguro1==True:
            v=(costo)*(1+nacionalizacion)
            s=v*0.01/mesesfact[0]
            costo_producto.append(s)
            total_otroopex.append(s)
            seguro.append("si")
            articulos.append("seguro plataforma")
            costos.append(s) #aca S debe llevar la formual 
            cantidades.append(1)
            tipo.append("seguro")
            c=componente_status[-1]
            componente.append(c)

            
        else:
            seguro.append("no")
            
        unitario=((1+nacionalizacion/100)*(1+iva/100))*((mesesfact[0]+mesesnofact)/mesesfact[0])
        unit=unitario*costo
        cuota = npf.pmt(tasa_financiacion/100, mesesfact[0], unit*-1*cantidad , 0,1)
        cuota_mensual_capex=round(cuota,3)
        total_capex.append(cuota_mensual_capex)
        articulos.append(item)
        cantidades.append(cantidad)
        costos.append(cuota_mensual_capex)
        tipo.append("capex")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class infocliente(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 
    def __init__(self,parent=None):
        super(infocliente, self).__init__(parent)
        loadUi('infocliente.ui', self)
        self.boton_regresar.clicked.connect(self.Variables)#boton abre personal
        self.boton_regresar_2.clicked.connect(self.Anadir)#boton abre personal
    
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show() 
    def Anadir(self):
         c=self.c1.toPlainText()
         crm.append(c)
         cliente1=self.c2.toPlainText()
         cliente.append(cliente1)
         t=self.comboBox.currentText()
         tipoofer.append(t)
         print("info cliente")
class Parametros(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 


 
    def __init__(self,parent=None):
        super(Parametros, self).__init__(parent)
        loadUi('Parametros.ui', self)
        self.boton_regresar.clicked.connect(self.Variables)#boton abre la pagina de atras variables
        self.boton.clicked.connect(self.anadir)#s
        
    
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
    def anadir(self):
        riesgo=self.p20.value()
        riesgototal.append(riesgo)
        margencapex=self.p14.value()
        tasa_capex.append(margencapex)
        tasaoex1t=self.p13.value()
        tasa_opex1t.append(tasaoex1t)
        trm=self.p19.value()
        ipcg=self.p15.value()
        ipc.append(ipcg)
        TRM.append(trm)
        meses=self.p21.value()
        mesesfact.append(meses)
        margenotro=self.p12.value()
        tasa_otroopex.append(margenotro)
        margenp=self.p10.value()
        tasa_personal.append(margenp)
        tasa_gnral=self.gg.value()
        tasa_financiacion_general.append(tasa_gnral)
        trm_serv=self.p24.value()
        TRM_SERV.append(trm_serv)
        
        impuestos=self.p15_2.value()
        dias_pago=self.p11.value()
        
        imprevistos=impuestos+((tasa_gnral + (dias_pago - 30))/30)
        imprevistos_1.append(imprevistos)

class DataFrameTable(tk.Frame):
    def __init__(self, parent=None, df=pd.DataFrame()):
        super().__init__()
        self.parent = parent
        self.pack(fill=tk.BOTH, expand=True)
        self.table = Table(
            self, dataframe=df,
            showtoolbar=False,
            showstatusbar=True,
            editable=True)
        self.table.show()       
class agregar(QtWidgets.QMainWindow, Ui_MainWindow): # aca entran los 4x elementos 

    def __init__(self,parent=None):
        super(agregar, self).__init__(parent)
        loadUi('agregar.ui', self)
        self.boton_regresar.clicked.connect(self.Variables)#boton abre personal
        self.boton_otroopex.clicked.connect(self.Otroopex)#boton abre personal
        self.boton_opex1t.clicked.connect(self.opex1t)#boton abre personal
        self.boton_capex.clicked.connect(self.Capex)#boton abre personal
        self.boton_personal.clicked.connect(self.Personal)#boton abre personal
        self.boton_agregados.clicked.connect(self.agregados)#boton abre personal
        self.boton_agregados_2.clicked.connect(self.eliminar)#boton abre personal
    
    def eliminar(self):
        if tipo[-1]=="personal":
            total_personal.pop(-1)
        if tipo[-1]=="capex":
            total_capex.pop(-1)
        if tipo[-1]=="otroopex":
            total_otroopex.pop(-1)
        if tipo[-1]=="opex1t":
            total_opex1t.pop(-1)
        if seguro[-1]=="si":
            total_otroopex.pop(-1)
        articulos.pop(-1)
        cantidades.pop(-1)
        costos.pop(-1)
        tipo.pop(-1)
        componente.pop(-1)
         
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
    def Personal(self): 
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Personal(self)
         otraventana.show()
         
    def Otroopex(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Otroopex(self)
         otraventana.show()
    def opex1t(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=opex1t(self)
         otraventana.show()
    def Capex(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Capex(self)
         otraventana.show()
    def agregar(self):
        self.hide()
        otraventana=agregar(self)
        otraventana.show()
    def agregados(self):
        print(len(costo_producto))
        print(len(costos))
        df = pd.DataFrame()
        df['Articulo']=articulos 
        df['Cantidad']=cantidades
        df['Costo']=costo_producto
        df['Costo mensual']=costos
        df['Tipo']=tipo
        df['Componente']=componente
        root = tk.Tk()
        table = DataFrameTable(root, df)
        root.mainloop()
class Tipooferta(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 
    def __init__(self,parent=None):
        super(Tipooferta, self).__init__(parent)
        loadUi('tipooferta.ui', self)
        self.boton_regresar.clicked.connect(self.Variables)#boton regresa
        self.boton_regresar_2.clicked.connect(self.abrir_adicion)#boton regresa
        self.boton_regresar_3.clicked.connect(self.abrir_nuevo)#boton regresa
        
    
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
    def abrir_adicion(self):
        os.startfile(archivo_adicion)
    def abrir_nuevo(self):
        os.startfile(archivo_nuevo)
class Variables(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de las variables 

    def __init__(self,parent=None):
        super(Variables, self).__init__(parent)
        loadUi('Variables.ui', self)
        self.boton_infocliente.clicked.connect(self.infocliente)# boton abre info cliete
        self.boton_parametros.clicked.connect(self.Parametros)# boton abre info parametros
        
        self.boton_salir.clicked.connect(self.close)# boton salir
        self.boton_agregar.clicked.connect(self.agregar)# boton salir 
        self.boton.clicked.connect(self.generar)# GENERAR
        self.boton_2.clicked.connect(self.generarofer)# GENERAR
        self.boton_3.clicked.connect(self.modificar)# GENERAR
        
    
    def generarofer(self):
         self.hide()
         otraventana=Tipooferta(self)
         otraventana.show()
    def infocliente(self):
         self.hide()
         otraventana=infocliente(self)
         otraventana.show()
    def Parametros(self):
         self.hide()
         otraventana=Parametros(self)
         otraventana.show()

    def agregar(self):
        self.hide()
        otraventana=Seleccion(self)
        otraventana.show()
        
    def modificar(self):
        self.hide()
        otraventana=Modificar(self)
        otraventana.show()
        
    def generar(self):
        wb = load_workbook(filesheet)

        # Seleccionamos el archivo
        sheet = wb.active
        g=tasa_personal[0]
        h=tasa_capex[0]
        j=tasa_opex1t[0]
        k=tasa_otroopex[0]
        l=tasa_financiacion_general[0]
        sheet['E18'] = g
        sheet['E19'] = h
        sheet['E20'] = j
        sheet['E21'] = k
        sheet['E22'] = l
        a=sum(total_capex)
        b=sum(total_opex1t)
        c=sum(total_otroopex)
        d=sum(total_personal)
        a1=(a/(1-(riesgototal[0]/100)-(tasa_capex[0]/100)-(imprevistos_1[0]/100)))*(1+(ipc[0]/100))
        b1=(b/(1-(riesgototal[0]/100)-(tasa_opex1t[0]/100)-  (imprevistos_1[0]/100)))*(1+(ipc[0]/100))
        c1=(c/(1-(riesgototal[0]/100)-(tasa_otroopex[0]/100)-(imprevistos_1[0]/100)))*(1+(ipc[0]/100))
        d1=(d/(1-(riesgototal[0]/100)-(tasa_personal[0]/100)-(imprevistos_1[0]/100)))*(1+(ipc[0]/100))
        
       

        # Ingresamos el valor 56 en la celda 'A1'
        sheet['A4'] = "VALOR Mensual"
        # Ingresamos el valor 1845 en la celda 'B3'
        sheet['B2']=today
        sheet['B3'] = "Capex"
        sheet['C3'] = "Otro opex"
        sheet['D3'] = "Opex 1t"
        sheet['E3'] = "Personal"
        sheet['B4'] = a1
        sheet['C4'] = c1
        sheet['D4'] = b1
        sheet['E4'] = d1
        sheet['A5'] = "TOTAL mensual"
        sheet['B5'] = (a1+b1+c1+d1)

        sheet['A6'] = "TOTAL PROYECTO"
        total1=(a1+b1+c1+d1)*mesesfact[0]
        sheet['B6'] = total1
        total.append(total1)
        # 
        sheet['A10'] = "Preventa"
        sheet['B10'] = preventa[0]
        sheet['A11'] = "Comercial"
        sheet['B11'] = comercial[0]
        sheet['B17'] = ipc[0]
        sheet['B18'] = riesgototal[0]
        sheet['B19'] = imprevistos_1[0]
        sheet['B20'] = crm[0]
        sheet['B21'] = linea[0]
        sheet['B12'] = cliente[0]
        sheet['B2'] = today
        
        
        
        
        ###################3
        df = pd.read_excel(archivo_reportes)
        nuevo_registro = {'comercial': comercial[-1], 'preventa': preventa[-1], 'cliente': cliente[-1],
                          'TOTAL': total1, 'CRM': crm[-1],'fecha':today,"tipo" : tipoofer[-1], 'Moneda':moneda[-1]}
        df = df.append(nuevo_registro, ignore_index=True)
        file_name = 'registros.xlsx'
        df.to_excel(file_name, index=False)
        #vamos a crear el segundo excel de la base
        df1 = pd.DataFrame()
        df1['Articulo']=articulos 
        df1['Cantidad']=cantidades
        df1['Costo']=costo_producto
        df1['Costo mensual']=costos
        df1['Tipo']=tipo
        nombre1=str(cliente[-1])
        file_name1 =nombre1 + '.xlsx'
        #####
        file_name2 =nombre1 + 'reporte' + '.xlsx'
        df1.to_excel(file_name1)
       # os.startfile(file_name1)
        # Guardamos el archivo con los cambios
        wb.save(file_name2)
        os.startfile(file_name2)   
class Seleccion(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 


    def __init__(self,parent=None):
        super(Seleccion, self).__init__(parent)
        loadUi('seleccion.ui', self)
        self.pushButton.clicked.connect(self.cop)#boton abre cop
        self.pushButton_2.clicked.connect(self.usd)#boton abre usd
    
    def cop(self):
         self.hide()
         otraventana=agregar(self)
         otraventana.show()
         moneda.append('COP')
    def usd(self):
         self.hide()
         otraventana=agregarUSD(self)
         otraventana.show()
         moneda.append('USD')
# para el costeo en USD
class agregarUSD(QtWidgets.QMainWindow, Ui_MainWindow): # aca entran los 4 elementos 

    def __init__(self,parent=None):
        super(agregarUSD, self).__init__(parent)
        loadUi('agregarusd.ui', self)
        self.boton_regresar.clicked.connect(self.Variables)#boton abre personal
        self.boton_otroopex.clicked.connect(self.Otroopex)#boton abre personal
        self.boton_opex1t.clicked.connect(self.opex1t)#boton abre personal
        self.boton_capex.clicked.connect(self.Capex)#boton abre personal
        self.boton_personal.clicked.connect(self.Personal)#boton abre personal
        self.boton_agregados.clicked.connect(self.agregados)#boton abre personal
        self.boton_agregados_2.clicked.connect(self.eliminar)#boton abre personal
    
    def eliminar(self):
        if tipo[-1]=="personal":
            total_personal.pop(-1)
        if tipo[-1]=="capex":
            total_capex.pop(-1)
        if tipo[-1]=="otroopex":
            total_otroopex.pop(-1)
        if tipo[-1]=="opex1t":
            total_opex1t.pop(-1)
        if seguro[-1]=="si":
            total_otroopex.pop(-1)
        articulos.pop(-1)
        cantidades.pop(-1)
        costos.pop(-1)
        tipo.pop(-1)
        componente.pop(-1)
         
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
    def Personal(self): 
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Personalusd(self)
         otraventana.show()
         
    def Otroopex(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Otroopexusd(self)
         otraventana.show()
    def opex1t(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=opex1tusd(self)
         otraventana.show()
    def Capex(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Capexusd(self)
         otraventana.show()
    def agregar(self):
        self.hide()
        otraventana=agregar(self)
        otraventana.show()
    def agregados(self):
            
        df = pd.DataFrame()
        df['Articulo']=articulos 
        df['Cantidad']=cantidades
        df['Costo']=costo_producto
        df['Costo']=costos
        df['Tipo']=tipo
        df['Componente']=componente
        root = tk.Tk()
        table = DataFrameTable(root, df)
        root.mainloop()
class Personalusd(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase del personal
 
    def __init__(self,parent=None):
        super(Personalusd, self).__init__(parent)
        loadUi('Personalusd.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre agregar
        self.boton_regresar1.clicked.connect(self.anadir)#boton abre agregar

    
    def agregar(self):
        self.hide()
        otraventana=agregarUSD(self)
        otraventana.show() 
    def anadir(self):
        rol=self.c1.toPlainText()
        perfil=self.c3.currentText()
        meses_ejecutados=self.c2.value()
        cantidad=self.c4.value()
        recargos=self.c5.value()
        ubicacion=self.c6.currentText()
        mesesno=self.c22.value()
        costo=1
        print("#####")
        print(ubicacion)
        print(perfil)
        if perfil=="Analista":
            costo=Analista1
        elif perfil=="Analista +":
            costo=Analista2
        elif perfil=="Analista ++":
            costo=Analista3
        elif perfil=="Profesional Junior":
            costo=Profesional1
        elif perfil=="Profesional Pleno":
            costo=Profesional2
        elif perfil=="Profesional Pleno +":
            costo=Profesional3
        elif perfil=="Profesional Especialista ":
            costo=Profesional4
        elif perfil=="Profesional Especialista +":
            costo=Profesional5
        elif perfil=="Profesional Especialista ++":
            costo=Profesional6
        elif perfil=="Profesional Senior ":
            costo=Profesional7
        elif perfil=="Profesional Senior +":
            costo=Profesional8
        elif perfil=="Experto ":
            costo=Experto
        if ubicacion=="En Sonda Pc":
            ubi=En1
        elif ubicacion=="En Cliente Pc":
            ubi=En2
        elif ubicacion=="SONDA Portatil":
            ubi=En3
        elif ubicacion=="Cliente Portatil":
            ubi=En4
        elif ubicacion=="En cliente sin equipo ":
            ubi=En5
        elif ubicacion=="En Sonda sin equipo":
            ubi=En6
        elif ubicacion=="Zona Franca RU1":
            ubi=En7
        elif ubicacion=="Zona Franca RU2":
            ubi=En8
        elif ubicacion=="Zona Franca RU3":
            ubi=En9
        elif ubicacion=="Field Services EUS":
            ubi=En10    
        P=(costo)+ubi
        costo_producto.append(P)
        cuota=P*cantidad*(meses_ejecutados+mesesno)/mesesfact[0]
        cuota_total_personal_=npf.pmt(tasa_financiacion_general[0]/100,mesesfact[0]/(meses_ejecutados+mesesno),-1*P*cantidad,0,1)
        cuota_total_personal_=cuota_total_personal_/TRM_SERV[-1]
        total_personal.append(cuota_total_personal_)
        articulos.append(rol)
        cantidades.append(cantidad)
        costos.append(cuota_total_personal_)
        tipo.append("personal")# aca empiezo a darle con los precios del personal 
        seguro.append("no")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class opex1tusd(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 
 
    def __init__(self,parent=None):
        super(opex1tusd, self).__init__(parent)
        loadUi('opex1tusd.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre personal
        self.anadi.clicked.connect(self.anadir)
    def agregar(self):
        self.hide()
        otraventana=agregarUSD(self)
        otraventana.show()
    def anadir(self):
    
        item=self.textEdit.toPlainText()
        veces_ejecutado=self.v.value()
        cantidad=self.c.value()
        nacio=self.n.value()
        costo=self.c1.value()
        costo_producto.append(costo)
        iva=self.i.value()
        unitario= costo*(1+nacio/100)*(1+iva/100)
        cuota_mensual_opex1t=npf.pmt(tasa_financiacion_general[0]/100,mesesfact[0]/veces_ejecutado,-1*unitario*cantidad,0,1)
        cop=self.checkBox.isChecked()
        if cop==True:
            cuota_mensual_opex1t=cuota_mensual_opex1t/TRM[0]
        total_opex1t.append(cuota_mensual_opex1t)
        articulos.append(item)
        cantidades.append(cantidad)
        costos.append(cuota_mensual_opex1t)
        tipo.append("opex1t")
        seguro.append("no")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class Otroopexusd(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de otro opex


 
    def __init__(self,parent=None):
        super(Otroopexusd, self).__init__(parent)
        loadUi('Otroopexusd.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre personal
        self.boton_anadir.clicked.connect(self.anadir)
        
    
    def agregar(self):
        self.hide()
        otraventana=agregarUSD(self)
        otraventana.show()
    def anadir(self):
    
        item=self.textEdit.toPlainText()
        meses_ejecutado=self.spinBox.value()
        cantidad=self.c.value()
        costo=self.c1.value()
        costo_producto.append(costo)
        iva=self.i.value()
        trm_servicio=self.c3.isChecked()
        cop=self.c3.isChecked()
        
        if cop==True:
            costo=costo/TRM[0]
        unit=costo*(1+iva/100)
        cuota_mensual_otropex=npf.pmt(tasa_financiacion_general[0]/100,mesesfact[0]/meses_ejecutado,-1*cantidad*unit,0,1)
        
        
        total_otroopex.append(cuota_mensual_otropex)
        articulos.append(item)
        cantidades.append(cantidad)
        costos.append(cuota_mensual_otropex)
        tipo.append("otroopex")
        seguro.append("no")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class Capexusd(QtWidgets.QMainWindow, Ui_MainWindow):# esta es la clase de Opex 1t 
         
    def __init__(self,parent=None):
        super(Capexusd, self).__init__(parent)
        loadUi('Capexusd.ui', self)
        self.boton_regresar.clicked.connect(self.agregar)#boton abre personal
        self.boton_anadir.clicked.connect(self.anadir)#boton añade
        
    
    def agregar(self):
        self.hide()
        otraventana=agregarUSD(self)
        otraventana.show()
    def anadir(self):
        

        # aca empiezo a darle con los precios 
        item=self.c1.toPlainText()
        empresa_fabricante=self.q1.toPlainText()
        mesesnofact=self.c2.value()
        nacionalizacion=self.c3.value()
        plataforma=self.c4.value()
        iva=self.c5.value()
        tasa_financiacion=self.c6.value()
        cantidad=self.c7.value()
        costo=self.c8.value()
        costo_producto.append(costo)
        dolar=self.c9.isChecked()
        seguro1=self.c10.isChecked()
        if dolar==True:
            costo=costo/TRM[0]
        if seguro1==True:
            v=(costo)*(1+nacionalizacion)
            s=v*0.01/mesesfact[0]
            costo_producto.append(s)
            total_otroopex.append(s)
            seguro.append("si")
            articulos.append("seguro plataforma")
            costos.append(s) #aca S debe llevar la formual 
            cantidades.append(1)
            tipo.append("seguro")
            c=componente_status[-1]
            componente.append(c)

            
        else:
            seguro.append("no")
            
        unitario=((1+nacionalizacion/100)*(1+iva/100))*((mesesfact[0]+mesesnofact)/mesesfact[0])
        unit=unitario*costo
        cuota = npf.pmt(tasa_financiacion/100, mesesfact[0], unit*-1*cantidad , 0,1)
        cuota_mensual_capex=round(cuota,3)
        total_capex.append(cuota_mensual_capex)
        articulos.append(item)
        cantidades.append(cantidad)
        costos.append(cuota_mensual_capex)
        tipo.append("capex")
        c=componente_status[-1]
        componente.append(c)
        componente_status.pop(-1)
class Modificar(QtWidgets.QMainWindow, Ui_MainWindow): # esta es la clase de Opex 1t 



 
    def __init__(self,parent=None):
        super(Modificar, self).__init__(parent)
        loadUi('modificar.ui', self)
        self.pushButton.clicked.connect(self.Variables)
        self.comboBox.addItems(list(reg1))
        self.pushButton_2.clicked.connect(self.abrir)
        self.pushButton_3.clicked.connect(self.buscar)
        
        
    def Variables(self):
         
         print(nm)
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
    def buscar(self):
        nombre_modificar=self.comboBox.currentText()
        nombre_modificar=nombre_modificar+'.xlsx'
        nm.append(nombre_modificar)

        file = open("file.txt", "w")
        file.write(nm[-1])
        file.close()
        print("guarda")
        print(nm[-1])
        print("####")
    def abrir(self):


        self.demo = DFEditor()
        self.demo.show()
        self.close()
class agregarM(QtWidgets.QMainWindow, Ui_MainWindow): # aca entran los 4 elementos 

    def __init__(self,parent=None):
        super(agregarM, self).__init__(parent)
        loadUi('agregar.ui', self)
        self.boton_regresar.clicked.connect(self.Variables)#boton abre personal
        self.boton_otroopex.clicked.connect(self.Otroopex)#boton abre personal
        self.boton_opex1t.clicked.connect(self.opex1t)#boton abre personal
        self.boton_capex.clicked.connect(self.Capex)#boton abre personal
        self.boton_personal.clicked.connect(self.Personal)#boton abre personal
        self.boton_agregados.clicked.connect(self.agregados)#boton abre personal
        self.boton_agregados_2.clicked.connect(self.eliminar)#boton abre personal
    
    def eliminar(self):
        if tipo[-1]=="personal":
            total_personal.pop(-1)
        if tipo[-1]=="capex":
            total_capex.pop(-1)
        if tipo[-1]=="otroopex":
            total_otroopex.pop(-1)
        if tipo[-1]=="opex1t":
            total_opex1t.pop(-1)
        if seguro[-1]=="si":
            total_otroopex.pop(-1)
        articulos.pop(-1)
        cantidades.pop(-1)
        costos.pop(-1)
        tipo.pop(-1)
        componente.pop(-1)
         
    def Variables(self):
         self.hide()
         otraventana=Variables(self)
         otraventana.show()
    def Personal(self): 
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Personal(self)
         otraventana.show()
         
    def Otroopex(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Otroopex(self)
         otraventana.show()
    def opex1t(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=opex1t(self)
         otraventana.show()
    def Capex(self):
         self.hide()
         status=self.comboBox.currentText()
         componente_status.append(status)
         otraventana=Capex(self)
         otraventana.show()
    def agregar(self):
        self.hide()
        otraventana=agregar(self)
        otraventana.show()
    def agregados(self):
        print(len(costo_producto))
        print(len(costos))
        df = pd.DataFrame()
        df['Articulo']=articulos 
        df['Cantidad']=cantidades
        df['Costo']=costo_producto
        df['Costo mensual']=costos
        df['Tipo']=tipo
        df['Componente']=componente
        root = tk.Tk()
        table = DataFrameTable(root, df)
        root.mainloop()
class FloatDelegate(QItemDelegate):
    def __init__(self, parent=None):
        super().__init__()

    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setValidator(QDoubleValidator())
        return editor

class TableWidget(QTableWidget):
    def __init__(self, df):
        super().__init__()
        self.df = df
        self.setStyleSheet('font-size: 20px;')

        # set table dimension
        nRows, nColumns = self.df.shape
        self.setColumnCount(nColumns)
        self.setRowCount(nRows)

        self.setHorizontalHeaderLabels(('Articulo', 'Cantidad','Costo','Costo Mensual','Tipo','Unitario'))

        self.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.setItemDelegateForColumn(1, FloatDelegate())

        # data insertion
        for i in range(self.rowCount()):
            for j in range(self.columnCount()):
                self.setItem(i, j, QTableWidgetItem(str(self.df.iloc[i, j])))

        self.cellChanged[int, int].connect(self.updateDF)   

    def updateDF(self, row, column):
        text = self.item(row, column).text()
        self.df.iloc[row, column] = text

class DFEditor(QWidget):
    
    # abrimos el excel como un dataframe


    f = open ('file.txt','r')
    mensaje = f.read()

    f.close()

    df = pd.read_excel(mensaje)
    df=df.round({'Costo': 1})
    df=df.round({'Costo mensual': 1})
    df['Costo']=df['Costo'].apply('{:,}'.format)
    #df['Costo mensual']=df['Costo mensual'].apply('{:,}'.format)

    G=df.to_numpy().transpose().tolist()

    lista1 =G[4]
    lista2=G[2]
    lista3=[]
    for i in range(0,len(lista1)):
        lista3.append(lista1[i]//lista2[i])


    data = {
        'Articulo': G[1],
        'Cantidad': G[2],
        'Costo': G[3],
        'Costo Mensual': G[4],
        'Tipo': G[5],
        'Unitario':lista3
        
    }

    df = pd.DataFrame(data)

    def __init__(self):
        super().__init__()
        self.resize(1500, 720)

        mainLayout = QVBoxLayout()

        self.table = TableWidget(DFEditor.df)
        mainLayout.addWidget(self.table)

        button_print = QPushButton('Guardar como nueva Version')
        button_print.setStyleSheet('font-size: 20px')
        button_print.clicked.connect(self.print_DF_Values)
        mainLayout.addWidget(button_print)

        button_export = QPushButton('regresar')
        button_export.setStyleSheet('font-size:20px')
        button_export.clicked.connect(self.export_to_csv)
        mainLayout.addWidget(button_export)     
        

        button_export = QPushButton('añadir articulos')
        button_export.setStyleSheet('font-size:20px')
        button_export.clicked.connect(self.agregar)
        mainLayout.addWidget(button_export)    

        self.setLayout(mainLayout)
    def agregar(self):
        self.hide()
        otraventana=agregarM(self)
        otraventana.show()
    def print_DF_Values(self):
        print(self.df)
        self.df=self.df
        a=0#capex
        b=0#opex1t
        c=0#otropex
        d=0#personal
        tipe=(self.df['Tipo'])
        unit=(self.df['Unitario'])
        cant=(self.df['Cantidad'])
        for i in range(len(self.df['Tipo'])):
            aaa=(self.df['Tipo'][i])
            if aaa=='capex':
                aab=float(unit[i])*float(cant[i])
                a+=aab
            if aaa=='opex1t':
                aab=float(unit[i])*float(cant[i])
                b+=aab
            if aaa=='otroopex':
                aab=float(unit[i])*float(cant[i])
                c+=aab
            if aaa=='persona':
                aab=float(unit[i])*float(cant[i])
                d+=aab

        # traigo los datos que neceisto del excel 
        
        print(a,b,c,d)
        #totales
        v=1
        f1=self.df
        f = open ('file.txt','r')
        mensaje = f.read()
        print(mensaje)
        f.close()
        mensaje=mensaje.replace(".xlsx","")
        print(mensaje)
        file_name1 =mensaje + 'v1'+ '.xlsx'
        mensaje1=mensaje

        f1.to_excel(file_name1)
        df = pd.read_excel(archivo_reportes)
        
        
        
        #abrir excel con el registro
        mensaje=mensaje+'reporte'+'.xlsx'
        excel_document = openpyxl.load_workbook(mensaje)
          
        sheet_obj = excel_document.active 
          
          
          
        cell_obj1 = sheet_obj.cell(row = 11, column = 2) 
        cell_obj2 = sheet_obj.cell(row = 10, column = 2) 
        cell_obj3 = sheet_obj.cell(row = 12, column = 2) 
        cell_obj4 = sheet_obj.cell(row = 6, column = 2) 
        cell_obj5 = sheet_obj.cell(row = 20, column = 2) 


        cell_obj7 = sheet_obj.cell(row = 19, column = 5) 
        cell_obj8 = sheet_obj.cell(row = 20, column = 5) 
        cell_obj9 = sheet_obj.cell(row = 21, column = 5) 
        cell_obj10 = sheet_obj.cell(row = 18, column = 5)
        
        cell_obj11 = sheet_obj.cell(row = 18, column = 2) 
        cell_obj12 = sheet_obj.cell(row = 19, column = 2) 
        cell_obj13 = sheet_obj.cell(row = 18, column = 5) 

        


        comercial= cell_obj1.value
        preventa=cell_obj2.value
        cliente=cell_obj3.value
        TOTAL=cell_obj4.value
        CRM=cell_obj5.value
        fecha=today
        tipoofer="new version"
        moneda="cop"
        tasa_capex=cell_obj7.value
        tasa_opex1t=cell_obj8.value
        tasa_otroopex=cell_obj9.value
        tasa_personal=cell_obj10.value
        riesgo=cell_obj11.value
        agregados=cell_obj12.value
        ipc=1
        mesesfact=36
        
        
        
        
        nuevo_registro = {'comercial': comercial, 'preventa': preventa, 'cliente': cliente,
                          'TOTAL': TOTAL, 'CRM': CRM,'fecha':today,"tipo" : tipoofer, 'Moneda':moneda}
        df = df.append(nuevo_registro, ignore_index=True)
        file_name = 'registros.xlsx'
        df.to_excel(file_name, index=False)
        a1=(a/(1-(riesgo/100)-(tasa_capex/100)-(agregados/100)))*(1+(ipc/100))
        b1=(b/(1-(riesgo/100)-(tasa_opex1t/100)-(agregados/100)))*(1+(ipc/100))
        c1=(c/(1-(riesgo/100)-(tasa_otroopex/100)-(agregados/100)))*(1+(ipc/100))
        d1=(d/(1-(riesgo/100)-(tasa_personal/100)-(agregados/100)))*(1+(ipc/100))
        # abro el arcivho ene xcel del reporte

        print(a1+b1+c1+d1)
        # voy a crear el nuevo archivo excel 
        wb = load_workbook(filesheet)
        sheet = wb.active
        sheet['A4'] = "VALOR Mensual"
        # Ingresamos el valor 1845 en la celda 'B3'
        sheet['B2']=today
        sheet['B3'] = "Capex"
        sheet['C3'] = "Otro opex"
        sheet['D3'] = "Opex 1t"
        sheet['E3'] = "Personal"
        sheet['B4'] = a1
        sheet['C4'] = c1
        sheet['D4'] = b1
        sheet['E4'] = d1
        sheet['A5'] = "TOTAL mensual"
        sheet['B5'] = (a1+b1+c1+d1)
        sheet['A6'] = "TOTAL PROYECTO"
        total2=(a1+b1+c1+d1)*mesesfact
        sheet['B6'] = total2
        sheet['A10'] = "Preventa"
        sheet['B10'] = preventa
        sheet['A11'] = "Comercial"
        sheet['B11'] = comercial
        sheet['B17'] = ipc
        sheet['B18'] = riesgo
        sheet['B19'] = agregados
        sheet['B20'] = CRM
        sheet['B12'] = cliente
        sheet['B2'] = today
        file_name2 = mensaje1+'reporte'+'V1'+'.xlsx'

       # os.startfile(file_name1)
        # Guardamos el archivo con los cambios
        wb.save(file_name2)
        os.startfile(file_name2) 
        

    def export_to_csv(self):
        self.hide()
        otraventana=Modificar(self)
        otraventana.show()
if __name__ == "__main__": # abre la pantalla principal

        app =  QtWidgets.QApplication(sys.argv)
        window = Ventana1()
        window.show()
        sys.exit(app.exec_())